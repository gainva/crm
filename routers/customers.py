from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import or_
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta
from database import get_db
from auth import get_current_user, require_platform
import models
import io
from fastapi.responses import StreamingResponse

router = APIRouter()


class CustomerCreate(BaseModel):
    name_cn: str
    name_en: str = ""
    address_cn: str = ""
    address_en: str = ""
    website: str = ""
    industry_categories: List[str] = []
    notes: str = ""
    dealer_id: Optional[int] = None


class CustomerUpdate(BaseModel):
    name_cn: Optional[str] = None
    name_en: Optional[str] = None
    address_cn: Optional[str] = None
    address_en: Optional[str] = None
    website: Optional[str] = None
    industry_categories: Optional[List[str]] = None
    notes: Optional[str] = None
    label_color: Optional[str] = None
    is_contracted: Optional[bool] = None


def _serialize(c: models.Customer) -> dict:
    now = datetime.utcnow()
    is_expired = (not c.is_contracted) and c.expires_at and c.expires_at < now
    return {
        "id": c.id,
        "dealer_id": c.dealer_id,
        "dealer_name": c.dealer.company_name_cn if c.dealer else "平台添加",
        "name_cn": c.name_cn,
        "name_en": c.name_en,
        "address_cn": c.address_cn,
        "address_en": c.address_en,
        "website": c.website,
        "industry_categories": c.industry_categories or [],
        "is_contracted": c.is_contracted,
        "contracted_at": c.contracted_at.isoformat() if c.contracted_at else None,
        "label_color": c.label_color,
        "expires_at": c.expires_at.isoformat() if c.expires_at else None,
        "is_expired": is_expired,
        "notes": c.notes,
        "added_by_platform": c.added_by_platform,
        "created_at": c.created_at.isoformat(),
        "updated_at": c.updated_at.isoformat() if c.updated_at else None,
    }


@router.get("")
def list_customers(
    search: str = Query(""),
    contracted: Optional[bool] = Query(None),
    dealer_id: Optional[int] = Query(None),
    skip: int = 0,
    limit: int = 50,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(models.Customer)

    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile:
            return []
        q = q.filter(models.Customer.dealer_id == profile.id)
    elif dealer_id:
        q = q.filter(models.Customer.dealer_id == dealer_id)

    if search:
        q = q.filter(or_(
            models.Customer.name_cn.contains(search),
            models.Customer.name_en.contains(search),
        ))
    if contracted is not None:
        q = q.filter(models.Customer.is_contracted == contracted)

    total = q.count()
    customers = q.order_by(models.Customer.updated_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [_serialize(c) for c in customers]}


@router.post("")
def create_customer(
    req: CustomerCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile:
            raise HTTPException(status_code=400, detail="No dealer profile")
        dealer_id = profile.id
        added_by_platform = False
        expires_at = datetime.utcnow() + timedelta(days=180)
    else:
        dealer_id = req.dealer_id
        added_by_platform = True
        expires_at = None

    customer = models.Customer(
        dealer_id=dealer_id,
        name_cn=req.name_cn,
        name_en=req.name_en,
        address_cn=req.address_cn,
        address_en=req.address_en,
        website=req.website,
        industry_categories=req.industry_categories,
        notes=req.notes,
        added_by_platform=added_by_platform,
        expires_at=expires_at,
    )
    db.add(customer)
    db.commit()
    db.refresh(customer)
    return _serialize(customer)


@router.get("/{customer_id}")
def get_customer(
    customer_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    c = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")
    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile or c.dealer_id != profile.id:
            raise HTTPException(status_code=403, detail="Access denied")
    return _serialize(c)


@router.put("/{customer_id}")
def update_customer(
    customer_id: int,
    req: CustomerUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    c = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")

    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile or c.dealer_id != profile.id:
            raise HTTPException(status_code=403, detail="Access denied")

    for field, value in req.model_dump(exclude_none=True).items():
        setattr(c, field, value)

    if req.is_contracted is True and not c.contracted_at:
        c.contracted_at = datetime.utcnow()
        c.expires_at = None
        c.label_color = "blue"

    c.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(c)
    return _serialize(c)


@router.delete("/{customer_id}")
def delete_customer(
    customer_id: int,
    current_user: models.User = Depends(require_platform),
    db: Session = Depends(get_db),
):
    c = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Customer not found")
    db.delete(c)
    db.commit()
    return {"message": "Customer deleted"}


@router.get("/export/excel")
def export_customers_excel(
    dealer_id: Optional[int] = Query(None),
    contracted: Optional[bool] = Query(None),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    q = db.query(models.Customer)
    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if profile:
            q = q.filter(models.Customer.dealer_id == profile.id)
    elif dealer_id:
        q = q.filter(models.Customer.dealer_id == dealer_id)
    if contracted is not None:
        q = q.filter(models.Customer.is_contracted == contracted)

    customers = q.order_by(models.Customer.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "客户列表"

    headers = ["客户名称(中文)", "客户名称(英文)", "地址(中文)", "地址(英文)", "网站",
               "行业分类", "是否成单", "有效期至", "代理商", "注册日期", "最后更新"]
    header_fill = PatternFill("solid", start_color="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    for row, c in enumerate(customers, 2):
        ws.cell(row=row, column=1, value=c.name_cn)
        ws.cell(row=row, column=2, value=c.name_en)
        ws.cell(row=row, column=3, value=c.address_cn)
        ws.cell(row=row, column=4, value=c.address_en)
        ws.cell(row=row, column=5, value=c.website)
        ws.cell(row=row, column=6, value=", ".join(c.industry_categories or []))
        ws.cell(row=row, column=7, value="是" if c.is_contracted else "否")
        ws.cell(row=row, column=8, value=c.expires_at.strftime("%Y-%m-%d") if c.expires_at else "永久有效")
        ws.cell(row=row, column=9, value=c.dealer.company_name_cn if c.dealer else "平台添加")
        ws.cell(row=row, column=10, value=c.created_at.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=11, value=c.updated_at.strftime("%Y-%m-%d") if c.updated_at else "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customers.xlsx"},
    )
