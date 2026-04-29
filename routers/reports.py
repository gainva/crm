from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta
from database import get_db
from auth import get_current_user, require_platform
import models
import io
from fastapi.responses import StreamingResponse

router = APIRouter()

VALIDITY_DAYS = 90  # 3 months


def next_reg_no(db: Session) -> str:
    """Generate a unique registration number, safe against gaps/deletions."""
    now = datetime.utcnow()
    prefix = f"NT{now.strftime('%y%m')}"
    # Find the highest existing sequential number for this month
    existing = db.query(models.Report.registration_no).filter(
        models.Report.registration_no.like(f"{prefix}-%")
    ).all()
    max_seq = 0
    for (reg_no,) in existing:
        try:
            seq = int(reg_no.split("-")[-1])
            if seq > max_seq:
                max_seq = seq
        except (ValueError, IndexError):
            pass
    return f"{prefix}-{str(max_seq + 1).zfill(4)}"


def _send_notification(db: Session, user_id: int, ntype: str, title: str, message: str, related_id: int):
    n = models.Notification(user_id=user_id, type=ntype, title=title, message=message, related_id=related_id)
    db.add(n)


def _serialize_report(r: models.Report, include_reviews: bool = False) -> dict:
    data = {
        "id": r.id,
        "registration_no": r.registration_no,
        "dealer_id": r.dealer_id,
        "dealer_name": r.dealer.company_name_cn if r.dealer else "",
        "customer_id": r.customer_id,
        "application_date": r.application_date.isoformat() if r.application_date else None,
        "applicant_name": r.applicant_name,
        "applicant_company": r.applicant_company,
        "contact_info": r.contact_info,
        "customer_name_cn": r.customer_name_cn,
        "customer_name_en": r.customer_name_en,
        "customer_address_cn": r.customer_address_cn,
        "customer_address_en": r.customer_address_en,
        "customer_website": r.customer_website,
        "industry_categories": r.industry_categories or [],
        "part_name": r.part_name,
        "final_product_use": r.final_product_use,
        "production_capacity": r.production_capacity,
        "similar_product_count": r.similar_product_count,
        "project_budget": r.project_budget,
        "delivery_deadline": r.delivery_deadline,
        "project_model": r.project_model,
        "estimated_quantity": r.estimated_quantity,
        "main_competitors": r.main_competitors,
        "investment_purpose": r.investment_purpose or [],
        "project_key_points": r.project_key_points or [],
        "sales_opinion": r.sales_opinion,
        "comm_type": r.comm_type,
        "comm_count": r.comm_count,
        "comm_person1_name": r.comm_person1_name,
        "comm_person1_title": r.comm_person1_title,
        "comm_person2_name": r.comm_person2_name,
        "comm_person2_title": r.comm_person2_title,
        "status": r.status,
        "current_review_level": r.current_review_level,
        "rejection_reason": r.rejection_reason,
        "extension_count": r.extension_count,
        "valid_until": r.valid_until.isoformat() if r.valid_until else None,
        "approved_at": r.approved_at.isoformat() if r.approved_at else None,
        "step1_date": r.step1_date.isoformat() if r.step1_date else None,
        "step2_date": r.step2_date.isoformat() if r.step2_date else None,
        "step3_date": r.step3_date.isoformat() if r.step3_date else None,
        "step4_date": r.step4_date.isoformat() if r.step4_date else None,
        "step5_date": r.step5_date.isoformat() if r.step5_date else None,
        "ai_score":             r.ai_score,
        "ai_score_details":     r.ai_score_details,
        "has_duplicate_warning":r.has_duplicate_warning,
        "duplicate_warnings":   r.duplicate_warnings,
        "created_at": r.created_at.isoformat(),
        "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        "attachments": [
            {
                "id": a.id,
                "file_type": a.file_type,
                "file_name": a.file_name,
                "original_name": a.original_name,
                "file_size": a.file_size,
                "created_at": a.created_at.isoformat(),
            }
            for a in r.attachments
        ],
    }
    if include_reviews:
        data["reviews"] = [
            {
                "id": rv.id,
                "level": rv.level,
                "action": rv.action,
                "reason": rv.reason,
                "reviewer": rv.reviewer.username if rv.reviewer else "",
                "created_at": rv.created_at.isoformat(),
            }
            for rv in r.reviews
        ]
    return data


class ReportCreate(BaseModel):
    applicant_name: str
    applicant_company: str
    contact_info: str = ""
    customer_name_cn: str
    customer_name_en: str = ""
    customer_address_cn: str = ""
    customer_address_en: str = ""
    customer_website: str = ""
    industry_categories: List[str] = []
    part_name: str = ""
    final_product_use: str = ""
    production_capacity: str = ""
    similar_product_count: str = ""
    project_budget: str = ""
    delivery_deadline: str = ""
    project_model: str = ""
    estimated_quantity: str = ""
    main_competitors: str = ""
    investment_purpose: List[str] = []
    project_key_points: List[str] = []
    sales_opinion: str = ""
    comm_type: str = ""
    comm_count: str = ""
    comm_person1_name: str = ""
    comm_person1_title: str = ""
    comm_person2_name: str = ""
    comm_person2_title: str = ""
    customer_id: Optional[int] = None


class ReportUpdate(BaseModel):
    applicant_name: Optional[str] = None
    applicant_company: Optional[str] = None
    contact_info: Optional[str] = None
    customer_name_cn: Optional[str] = None
    customer_name_en: Optional[str] = None
    customer_address_cn: Optional[str] = None
    customer_address_en: Optional[str] = None
    customer_website: Optional[str] = None
    industry_categories: Optional[List[str]] = None
    part_name: Optional[str] = None
    final_product_use: Optional[str] = None
    production_capacity: Optional[str] = None
    similar_product_count: Optional[str] = None
    project_budget: Optional[str] = None
    delivery_deadline: Optional[str] = None
    project_model: Optional[str] = None
    estimated_quantity: Optional[str] = None
    main_competitors: Optional[str] = None
    investment_purpose: Optional[List[str]] = None
    project_key_points: Optional[List[str]] = None
    sales_opinion: Optional[str] = None
    comm_type: Optional[str] = None
    comm_count: Optional[str] = None
    comm_person1_name: Optional[str] = None
    comm_person1_title: Optional[str] = None
    comm_person2_name: Optional[str] = None
    comm_person2_title: Optional[str] = None


class ReviewAction(BaseModel):
    action: str  # approve | reject
    reason: str = ""


class ProgressUpdate(BaseModel):
    step: int  # 1-5
    date: Optional[str] = None  # ISO date string or null to clear


@router.get("")
def list_reports(
    search: str = Query(""),
    status: str = Query(""),
    dealer_id: Optional[int] = Query(None),
    skip: int = 0,
    limit: int = 50,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(models.Report)

    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile:
            return {"total": 0, "items": []}
        q = q.filter(models.Report.dealer_id == profile.id)
    elif current_user.role == "platform_staff":
        emp = current_user.employee
        if emp and not emp.can_view_all:
            q = q.filter(models.Report.current_review_level == emp.review_level)
    elif dealer_id:
        q = q.filter(models.Report.dealer_id == dealer_id)

    if search:
        q = q.filter(or_(
            models.Report.registration_no.contains(search),
            models.Report.customer_name_cn.contains(search),
            models.Report.applicant_company.contains(search),
        ))
    if status:
        q = q.filter(models.Report.status == status)

    total = q.count()
    reports = q.order_by(models.Report.updated_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [_serialize_report(r) for r in reports]}


@router.post("")
def create_report(
    req: ReportCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("dealer", "admin", "platform_staff"):
        raise HTTPException(status_code=403, detail="Forbidden")

    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile:
            raise HTTPException(status_code=400, detail="No dealer profile")
        dealer_id = profile.id
    else:
        dealer_id = req.customer_id  # platform can create on behalf

    report = models.Report(
        dealer_id=dealer_id if current_user.role == "dealer" else None,
        registration_no=next_reg_no(db),
        applicant_name=req.applicant_name,
        applicant_company=req.applicant_company,
        contact_info=req.contact_info,
        customer_name_cn=req.customer_name_cn,
        customer_name_en=req.customer_name_en,
        customer_address_cn=req.customer_address_cn,
        customer_address_en=req.customer_address_en,
        customer_website=req.customer_website,
        industry_categories=req.industry_categories,
        part_name=req.part_name,
        final_product_use=req.final_product_use,
        production_capacity=req.production_capacity,
        similar_product_count=req.similar_product_count,
        project_budget=req.project_budget,
        delivery_deadline=req.delivery_deadline,
        project_model=req.project_model,
        estimated_quantity=req.estimated_quantity,
        main_competitors=req.main_competitors,
        investment_purpose=req.investment_purpose,
        project_key_points=req.project_key_points,
        sales_opinion=req.sales_opinion,
        comm_type=req.comm_type,
        comm_count=req.comm_count,
        comm_person1_name=req.comm_person1_name,
        comm_person1_title=req.comm_person1_title,
        comm_person2_name=req.comm_person2_name,
        comm_person2_title=req.comm_person2_title,
        customer_id=req.customer_id,
        status="draft",
    )
    if current_user.role == "dealer":
        report.dealer_id = current_user.dealer_profile.id

    db.add(report)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        # Retry with a fresh registration number (race condition safety)
        report.registration_no = next_reg_no(db)
        db.add(report)
        db.commit()
    db.refresh(report)
    return _serialize_report(report)


@router.get("/{report_id}")
def get_report(
    report_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    r = db.query(models.Report).filter(models.Report.id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Report not found")
    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile or r.dealer_id != profile.id:
            raise HTTPException(status_code=403, detail="Access denied")
    return _serialize_report(r, include_reviews=True)


@router.put("/{report_id}")
def update_report(
    report_id: int,
    req: ReportUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    r = db.query(models.Report).filter(models.Report.id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Report not found")

    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile or r.dealer_id != profile.id:
            raise HTTPException(status_code=403, detail="Access denied")
        if r.status not in ("draft", "rejected"):
            raise HTTPException(status_code=400, detail="Cannot edit report in current status")

    for field, value in req.model_dump(exclude_none=True).items():
        setattr(r, field, value)

    r.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(r)
    return _serialize_report(r)


@router.post("/{report_id}/submit")
def submit_report(
    report_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from ai_utils import calculate_ai_score, detect_duplicates

    r = db.query(models.Report).filter(models.Report.id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Report not found")

    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile or r.dealer_id != profile.id:
            raise HTTPException(status_code=403, detail="Access denied")

    if r.status not in ("draft", "rejected"):
        raise HTTPException(status_code=400, detail="Report cannot be submitted in current status")

    if not r.customer_name_cn or not r.applicant_name or not r.applicant_company:
        raise HTTPException(status_code=400, detail="Required fields missing")

    # ── AI Lead Scoring ────────────────────────────────────────────────────
    try:
        score, score_details = calculate_ai_score(r, db)
        r.ai_score = score
        r.ai_score_details = score_details
    except Exception:
        r.ai_score = None
        r.ai_score_details = None

    # ── Duplicate Detection ────────────────────────────────────────────────
    try:
        dupes = detect_duplicates(r, db)
        r.duplicate_warnings = dupes if dupes else None
        r.has_duplicate_warning = bool(dupes and any(d["level"] == "high" for d in dupes))
    except Exception:
        r.duplicate_warnings = None
        r.has_duplicate_warning = False

    r.status = "pending_l1"
    r.current_review_level = 1
    r.rejection_reason = None
    r.updated_at = datetime.utcnow()
    db.commit()

    # ── Notify Level 1 reviewers ───────────────────────────────────────────
    l1_reviewers = db.query(models.Employee).filter(
        models.Employee.review_level == 1,
        models.Employee.can_review == True,
    ).all()
    dupe_note = "⚠️ 系统检测到疑似重复客户记录，请重点核查！" if r.has_duplicate_warning else ""
    for emp in l1_reviewers:
        _send_notification(
            db, emp.user_id, "report_submitted",
            f"新报备待审核: {r.registration_no}（AI评分: {r.ai_score or '—'}）",
            f"代理商 {r.dealer.company_name_cn if r.dealer else ''} 提交了新报备。{dupe_note}",
            r.id,
        )
    db.commit()
    return _serialize_report(r)


@router.post("/{report_id}/review")
def review_report(
    report_id: int,
    req: ReviewAction,
    current_user: models.User = Depends(require_platform),
    db: Session = Depends(get_db),
):
    r = db.query(models.Report).filter(models.Report.id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Report not found")

    emp = current_user.employee
    if not emp:
        raise HTTPException(status_code=403, detail="No employee profile")
    if not emp.can_review:
        raise HTTPException(status_code=403, detail="No review permission")

    level_map = {"pending_l1": 1, "pending_l2": 2}
    if r.status not in level_map:
        raise HTTPException(status_code=400, detail="Report not pending review")

    expected_level = level_map[r.status]
    if emp.review_level != expected_level and current_user.role != "admin":
        raise HTTPException(status_code=403, detail=f"This report requires Level {expected_level} reviewer")

    if req.action not in ("approve", "reject"):
        raise HTTPException(status_code=400, detail="Action must be approve or reject")

    review = models.ReportReview(
        report_id=r.id,
        reviewer_id=current_user.id,
        level=expected_level,
        action=req.action,
        reason=req.reason,
    )
    db.add(review)

    if req.action == "approve":
        if expected_level == 1:
            r.status = "pending_l2"
            r.current_review_level = 2
            # Notify level 2
            l2 = db.query(models.Employee).filter(models.Employee.review_level == 2, models.Employee.can_review == True).all()
            for emp2 in l2:
                _send_notification(db, emp2.user_id, "report_submitted",
                                   f"报备待二级审核: {r.registration_no}", f"一级审核通过，请进行二级审核。", r.id)
        elif expected_level == 2:
            r.status = "approved"
            r.current_review_level = 0
            r.approved_at = datetime.utcnow()
            r.valid_until = datetime.utcnow() + timedelta(days=VALIDITY_DAYS)
            if r.dealer:
                _send_notification(db, r.dealer.user_id, "report_approved",
                                   f"报备已批准: {r.registration_no}",
                                   f"您的报备申请已获批准，有效期至 {r.valid_until.strftime('%Y-%m-%d')}。", r.id)
    else:  # reject
        if not req.reason:
            raise HTTPException(status_code=400, detail="Rejection reason required")
        r.status = "rejected"
        r.rejection_reason = req.reason
        r.current_review_level = 0
        if r.dealer:
            _send_notification(db, r.dealer.user_id, "report_rejected",
                               f"报备已驳回: {r.registration_no}",
                               f"您的报备申请被驳回，原因: {req.reason}", r.id)

    r.updated_at = datetime.utcnow()
    db.commit()
    return _serialize_report(r, include_reviews=True)


@router.post("/{report_id}/extend")
def extend_report(
    report_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    r = db.query(models.Report).filter(models.Report.id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Report not found")

    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile or r.dealer_id != profile.id:
            raise HTTPException(status_code=403, detail="Access denied")

    if r.status != "approved":
        raise HTTPException(status_code=400, detail="Only approved reports can be extended")
    if r.extension_count >= 2:
        raise HTTPException(status_code=400, detail="Maximum extensions reached (max 2 × 3 months)")

    r.extension_count += 1
    r.valid_until = (r.valid_until or datetime.utcnow()) + timedelta(days=90)
    r.updated_at = datetime.utcnow()
    db.commit()
    return _serialize_report(r)


@router.post("/{report_id}/mark-contracted")
def mark_contracted(
    report_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    r = db.query(models.Report).filter(models.Report.id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Report not found")

    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile or r.dealer_id != profile.id:
            raise HTTPException(status_code=403, detail="Access denied")

    r.status = "contracted"
    r.step5_date = datetime.utcnow()
    r.valid_until = None  # Permanent
    r.updated_at = datetime.utcnow()

    # Mark associated customer as contracted
    if r.customer_id:
        c = db.query(models.Customer).filter(models.Customer.id == r.customer_id).first()
        if c:
            c.is_contracted = True
            c.contracted_at = datetime.utcnow()
            c.expires_at = None
            c.label_color = "blue"

    db.commit()
    return _serialize_report(r)


@router.put("/{report_id}/progress")
def update_progress(
    report_id: int,
    req: ProgressUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    r = db.query(models.Report).filter(models.Report.id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Report not found")

    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile or r.dealer_id != profile.id:
            raise HTTPException(status_code=403, detail="Access denied")

    if req.step < 1 or req.step > 5:
        raise HTTPException(status_code=400, detail="Step must be 1-5")

    date_val = datetime.fromisoformat(req.date) if req.date else None
    setattr(r, f"step{req.step}_date", date_val)
    r.updated_at = datetime.utcnow()
    db.commit()
    return _serialize_report(r)


@router.delete("/{report_id}")
def delete_report(
    report_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    r = db.query(models.Report).filter(models.Report.id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Report not found")

    if current_user.role == "dealer":
        profile = current_user.dealer_profile
        if not profile or r.dealer_id != profile.id:
            raise HTTPException(status_code=403, detail="Access denied")
        if r.status not in ("draft", "rejected"):
            raise HTTPException(status_code=400, detail="Cannot delete submitted reports")

    for a in r.attachments:
        import os
        if os.path.exists(a.file_path):
            os.remove(a.file_path)
        db.delete(a)
    for rv in r.reviews:
        db.delete(rv)
    db.delete(r)
    db.commit()
    return {"message": "Report deleted"}


@router.get("/export/excel")
def export_reports_excel(
    status: str = Query(""),
    dealer_id: Optional[int] = Query(None),
    current_user: models.User = Depends(require_platform),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    q = db.query(models.Report)
    if dealer_id:
        q = q.filter(models.Report.dealer_id == dealer_id)
    if status:
        q = q.filter(models.Report.status == status)
    reports = q.order_by(models.Report.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "报备列表"

    headers = ["登録番号", "申请公司", "申请人", "客户名称(中文)", "零件名称", "项目预算",
               "状态", "有效期至", "审批日期", "登记日期"]
    hfill = PatternFill("solid", start_color="1E3A5F")
    hfont = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    status_map = {
        "draft": "草稿", "pending_l1": "待一审", "pending_l2": "待二审",
        "approved": "已批准", "rejected": "已驳回",
        "expired": "已过期", "contracted": "已成单",
    }
    for row, r in enumerate(reports, 2):
        ws.cell(row=row, column=1, value=r.registration_no)
        ws.cell(row=row, column=2, value=r.applicant_company)
        ws.cell(row=row, column=3, value=r.applicant_name)
        ws.cell(row=row, column=4, value=r.customer_name_cn)
        ws.cell(row=row, column=5, value=r.part_name)
        ws.cell(row=row, column=6, value=r.project_budget)
        ws.cell(row=row, column=7, value=status_map.get(r.status, r.status))
        ws.cell(row=row, column=8, value=r.valid_until.strftime("%Y-%m-%d") if r.valid_until else "")
        ws.cell(row=row, column=9, value=r.approved_at.strftime("%Y-%m-%d") if r.approved_at else "")
        ws.cell(row=row, column=10, value=r.created_at.strftime("%Y-%m-%d"))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reports.xlsx"},
    )
